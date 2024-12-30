from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import requests
from bs4 import BeautifulSoup
import openpyxl 

# Bağlantı sorunları için hata yönetimi ve URL'den içerik alma fonksiyonu
def url_icerik_getir(url):
    try:
        response = requests.get(url)
        response.raise_for_status()
        print("Sayfa başarıyla indirildi.")
        return response.content
    except requests.exceptions.RequestException as e:
        print(f"URL'ye erişim başarısız: {e}")
        return None

url = "https://ebs.kocaelisaglik.edu.tr/Pages/CourseDetail.aspx?lang=tr-TR&academicYear=2024&facultyId=5&programId=1&menuType=course&catalogId=2227"
html_content = url_icerik_getir(url)
if not html_content:
    exit()

# 2. HTML'i ayrıştırma ve arama 
soup = BeautifulSoup(html_content, "html.parser")

# 3. Hedef Tabloyu Bulma
ders_table = soup.find("table", id="Content_Content_LearningOutcomes_gridLearningOutComes_DXMainTable")
program_table = soup.find("table", id="Content_Content_DersinCiktilaraKatkisi_gridDersinCiktilaraEtkisi_DXMainTable")

# 4. Ders Çıktılarını İşleme
ders_ciktilari = []
if ders_table:
    rows = ders_table.find_all(
        "tr", id=lambda x: x and x.startswith("Content_Content_LearningOutcomes_gridLearningOutComes_DXDataRow")
    )
    for row in rows:
        columns = row.find_all("td")
        if len(columns) >= 2:
            sayi = columns[0].text.strip()
            cikti = columns[1].text.strip()
            ders_ciktilari.append({"Sayı": sayi, "Öğrenme Çıktısı": cikti})
else:
    print("Ders çıktıları tablosu bulunamadı!")

# 5. Program Çıktılarını İşleme
program_ciktilari = []
if program_table:
    rows = program_table.find_all(
        "tr", id=lambda x: x and x.startswith("Content_Content_DersinCiktilaraKatkisi_gridDersinCiktilaraEtkisi_DXDataRow")
    )
    for row in rows:
        columns = row.find_all("td")
        if len(columns) >= 2:
            sira_no = columns[0].text.strip()
            program_cikti = columns[1].text.strip()
            program_ciktilari.append({"Sıra No": sira_no, "Program Çıktısı": program_cikti})
else:
    print("Program çıktıları tablosu bulunamadı!")

# 6. Veriyi Excel'e Kaydetme
if ders_ciktilari:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ders Çıktıları"

    # Başlıkları ekle
    ws.append(["Sayı", "Öğrenme Çıktısı"])

    # Veriyi ekle
    for cikti in ders_ciktilari:
        ws.append([cikti["Sayı"], cikti["Öğrenme Çıktısı"]])

    wb.save("Ders_Ciktilari.xlsx")
    print("Ders çıktıları başarıyla kaydedildi: Ders_Ciktilari.xlsx")
else:
    print("Ders çıktıları kaydedilemedi çünkü veri bulunamadı.")

if program_ciktilari:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Program Çıktıları"

    # Başlıkları ekle
    headers = ["Sıra No", "Program Çıktısı"] + [f"Katkı {i+1}" for i in range(len(program_ciktilari[0]) - 2)]
    ws.append(headers)

    # Veriyi ekle
    for cikti in program_ciktilari:
        ws.append(list(cikti.values()))

    wb.save("Program_Ciktilari_Katkilar.xlsx")
    print("Program çıktıları başarıyla kaydedildi: Program_Ciktilari_Katkilar.xlsx")
else:
    print("Program çıktıları kaydedilemedi çünkü veri bulunamadı.")
file_path = r'C:/Users/PC/Desktop/kod/Python/Değerlendirme test çıktıları.xlsx'  #Excel dosyasını aç veya oluştur
try:
    ç_sayfası = load_workbook(file_path)
except FileNotFoundError:
    from openpyxl import Workbook
    ç_sayfası = Workbook()
    ç_sayfası.active.title = "Sayfa1"
    ç_sayfası.save(file_path)
sayfa = ç_sayfası.active  #İlk sayfa seçilir 

sayfa.merge_cells('C2:G2') #merge_cells hücreleri birleştirmek için bir fonksiyon 'C2' ile 'G2' arasındaki hücreleri birleştir
sayfa['C2'] = "Ders Çıktısı" #Metni 'C2' hücresine yerleştir

başlık = ["1. Başlık", "2. Başlık", "3. Başlık", "4. Başlık", "5. Başlık", "İlişki Değeri"] #Başlıklar
sayfa['B2'] = "Tablo 1"
sayfa['B3'] = "Program çıktısı"
sayfa['C3'] = başlık[0]  
sayfa['D3'] = başlık[1]  
sayfa['E3'] = başlık[2]  
sayfa['F3'] = başlık[3]  
sayfa['G3'] = başlık[4]  
sayfa['H3'] = başlık[5]  
programlar = ["Program 1", "Program 2", "Program 3", "Program 4", "Program 5", "Program 6", "Program 7", "Program 8", "Program 9", "Program 10"]

for i, program in enumerate(programlar, start=4):
    sayfa[f'B{i}'] = program
def add_row1(data11, start_row=4, end_row=13): #Yeni veri eklemek için fonksiyon
    for row_num in range(start_row, end_row + 1): #Başlangıç satırından itibaren her satıra veri ekle
        for col_num, value in enumerate(data11[row_num - start_row], start=3): #3. sütun ile başla
            sayfa.cell(row=row_num, column=col_num, value=value)
        column_count = len(data11[row_num - start_row]) #Veri eklediğimiz SÜTUN sayısını dinamik olarak tespit eder (C, D, E, F, G gibi)
        değerler = [sayfa.cell(row=row_num, column=col_num).value for col_num in range(3, 3 + column_count)] #Sütunlardan gelen değerleri al
        toplam = sum(değerler) #Toplam değeri hesapla
        sonuç = toplam / column_count #Verilen sütun sayısına göre bölme işlemi ortalamayı hesaplar
        sayfa.cell(row=row_num, column=8, value=sonuç) #Sonucu aynı satırda H hücresine yazar

#--------------------------------------------------------------------Tablo 1'i burdan elle değiştiriyoruz
data1 = [
    [1, 1, 1, 0.5, 0],    #Tablo 1'deki progrlamların değerleri
    [0, 1, 1, 0.7, 0.9],  
    [0, 0, 0.3, 1, 1],    
    [0, 0, 1, 0.6, 0.4],  
    [0, 0, 0, 0, 1],     
    [1, 1, 0, 0, 0],      
    [0.7, 0.8, 0, 0, 0.5],
    [0.1, 0.3, 1, 0.4, 0],
    [0.6, 1, 0, 0, 0.7],  
    [1, 0.4, 0, 0, 0.8]   
]

add_row1(data1, start_row=4, end_row=13) #Verileri ekle
sayfa.merge_cells('J1:O1') #'J1' ile 'O1'arasındaki hücreleri birleştir
sayfa['J1'] = "Değerlendirmeler" #'J1' hücresine yaz
sayfa['O3'] = "Toplam"
sayfa['J2'] = "Tablo 2"
başlık = ["Ders Çıktısı", "Ödev", "Quiz", "Vize", "Final"] #Sırasıyla başlıkları yazalım

for col_num, header in enumerate(başlık, start=10):# 10, J sütunu için
    sayfa.cell(row=3, column=col_num, value=header)

numbers_str = ["10", "20", "30", "40"]  #Burada string olarak ödev, quiz, vize ve final için değer
numbers = [int(num) for num in numbers_str]

for col_num, number in enumerate(numbers, start=11):#11, K sütunu için
    sayfa.cell(row=2, column=col_num, value=number)

subjects = ["1. Başlık", "2. Başlık", "3. Başlık", "4. Başlık", "5. Başlık"]
for row_num, subject in enumerate(subjects, start=4):#4, J4 hücresinden başlamak için
    sayfa.cell(row=row_num, column=10, value=subject)#10, J sütunu için

# Yeni veri eklemek için fonksiyon
def add_row2(data11, start_row=4, end_row=8):
    for row_num in range(start_row, end_row + 1): #Başlangıç satırından itibaren, her satıra veri ekle
        for col_num, value in enumerate(data11[row_num - start_row], start=11):#K sütunu (11. sütun) ile başla
            sayfa.cell(row=row_num, column=col_num, value=value) #Veri eklediğimiz sütun sayısını dinamik olarak tespit eder (K, L, M, N, O gibi)
        column_count = len(data11[row_num - start_row])
        değerler = [sayfa.cell(row=row_num, column=col_num).value for col_num in range(11, 11 + column_count)] #Sütunlardan gelen değerleri alır
        toplam = sum(değerler) #Toplam değeri hesaplar
        sayfa.cell(row=row_num, column=15, value=toplam) #Toplamı O sütununa yazar
    
#Satır ekleyelim--------------------------------------------------------------Tablo 2 yi burdan elle değiştirebiliriz.
data11 = [
    [0, 0, 1, 1],  
    [1, 1, 1, 0], 
    [1, 1, 1, 1], 
    [0, 1, 1, 1],  
    [0, 0, 0, 1]  
]
# Veri ekleme işlemini başlat
add_row2(data11, start_row=4, end_row=8)


sayfa['Q2'] = "Tablo 3"
sayfa.merge_cells('R2:V2') #'R2' ile 'V2' arasındaki hücreleri birleştir
sayfa['R2'] = "Ağırlıklı Değerlendirme" #Metni 'R2' hücresine yerleştir
subjects = ["1. Başlık", "2. Başlık", "3. Başlık", "4. Başlık", "5. Başlık"] 

for row_num, subject in enumerate(subjects, start=4):#4, Q4 hücresinden başlamak için
    sayfa.cell(row=row_num, column=17, value=subject) #17, Q sütunu için
    
başlık = ["Ders Çıktısı", "Ödev", "Quiz", "Vize", "Final", "Toplam"]  # Yeni başlıklar

for col_num, header in enumerate(başlık, start=17): #17, Q sütunu için
    sayfa.cell(row=3, column=col_num, value=header) #Q3, R3, S3, T3, U3 için başlıkları yerleştiriyoruz

def add_row3(data11, numbers, start_row=4, end_row=8): # Yeni veri eklemek için fonksiyon
    for row_num in range(start_row, end_row + 1): #Başlangıç satırından itibaren her satıra veri ekle
        for col_num, value in enumerate(data11[row_num - start_row], start=11):#K sütunu (11. sütun) ile başla
            sayfa.cell(row=row_num, column=col_num, value=value)
        column_count = len(data11[row_num - start_row])#Veri eklediğimiz sütun sayısını dinamik olarak tespit et (K, L, M, N gibi)
        değerler = [sayfa.cell(row=row_num, column=col_num).value for col_num in range(11, 11 + column_count)]#Sütunlardan gelen değerleri al
        sonuçlar = [] #Her hücredeki değeri sayılar ile çarpar ve 100'e böler
        for i, value in enumerate(değerler):
            sonuç = (value * numbers[i]) / 100
            sonuçlar.append(sonuç)
        # Sonuçları R, S, T, U, sütunlarına yaz 
        for i, sonuç in enumerate(sonuçlar, start=18):#18 = R sütunu
            sayfa.cell(row=row_num, column=i, value=sonuç)
        toplam = sum(sonuçlar) #Sonuçları toplaar ve toplamı sağdaki hücreye yazar
        sayfa.cell(row=row_num, column=22, value=toplam) #V sütununa toplamı yaz
# Fonksiyonu çağırarak verileri ve hesaplanmış sonuçları ekleyelim
add_row3(data11, numbers)

sayfa['B17'] = "Not Tablosu"
sayfa.merge_cells('C17:G17')
sayfa['C17'] = "Notlar"

başlık2 = ["Öğrenci", "Ödev", "Quiz", "Vize", "Final", "Ortalama"]  # Yeni başlıklar

for col_num, header in enumerate(başlık2, start=2):#17, Q sütunu için
    sayfa.cell(row=18, column=col_num, value=header)#Q3, R3, S3, T3, U3 için başlıkları yerleştiriyoruz

not_yukle_path = r'C:/Users/PC/Desktop/kod/Python/NotYukle.xlsx' #"not yükle.xlsx" dosyasının yolu
not_yukle_ç_sayfası = load_workbook(not_yukle_path) #"not yükle.xlsx" dosyasını aç
not_yukle_sayfa = not_yukle_ç_sayfası.active
start_row = 19 #Verilerin yazılacağı ilk satır

# "NotYukle.xlsx" dosyasındaki tüm verileri alır
# Buradaki veriler A, B, C, D ve E sütunlarında olacak (Öğrenci Numarası, Ödev, Quiz, Vize, Final)
for row in range(2, not_yukle_sayfa.max_row + 1):  # 2. satırdan başlayarak tüm satırları işle
    ogr_numara = not_yukle_sayfa.cell(row=row, column=1).value  # Öğrenci numarası
    odev = not_yukle_sayfa.cell(row=row, column=2).value        # Ödev notu
    quiz = not_yukle_sayfa.cell(row=row, column=3).value        # Quiz notu
    vize = not_yukle_sayfa.cell(row=row, column=4).value        # Vize notu
    final = not_yukle_sayfa.cell(row=row, column=5).value       # Final notu
    # Exel dosyasına B19 hücresinden itibaren öğrenci numarasını ve notları yerleştirir
    sayfa.cell(row=start_row, column=2, value=ogr_numara)  # B sütununa (Öğrenci numarası)
    sayfa.cell(row=start_row, column=3, value=odev)        # C sütununa (Ödev notu)
    sayfa.cell(row=start_row, column=4, value=quiz)        # D sütununa (Quiz notu)
    sayfa.cell(row=start_row, column=5, value=vize)        # E sütununa (Vize notu)
    sayfa.cell(row=start_row, column=6, value=final)       # F sütununa (Final notu)
    start_row += 1 # Bir sonraki satıra geç

k2 = sayfa['K2'].value  # K2 hücresindeki değer
l2 = sayfa['L2'].value  # L2 hücresindeki değer
m2 = sayfa['M2'].value  # M2 hücresindeki değer
n2 = sayfa['N2'].value  # N2 hücresindeki değer
start_row = 19 #19. satırdan başlayarak, son satıra kadar aynı işlemi yapacağız
end_row = sayfa.max_row  #Son satırı bulur

for row in range(start_row, end_row + 1): #19'dan son satıra kadar
    c = sayfa.cell(row=row, column=3).value  # C sütunu (Ödev)
    d = sayfa.cell(row=row, column=4).value  # D sütunu (Quiz)
    e = sayfa.cell(row=row, column=5).value  # E sütunu (Vize)
    f = sayfa.cell(row=row, column=6).value  # F sütunu (Final)
    if c is not None and d is not None and e is not None and f is not None:
        sonuç = (k2 * c / 100) + (l2 * d / 100) + (m2 * e / 100) + (n2 * f / 100) #Sonuçları hesaplar
        sayfa.cell(row=row, column=7, value=sonuç).font = Font(color="FF0000") #Sonuçları kırmızı renk yazar

#Tablo yapımı               
ogrenci_numaralari = []
row = 19  #B19 hücresinden başlar
# Öğrenci numaralarını alalım
while sayfa.cell(row=row, column=2).value is not None: #Öğrenci numarası varsa
    ogrenci_numaralari.append(sayfa.cell(row=row, column=2).value)
    row += 1  # Sonraki satıra geç
for i, ogrenci_numarasi in enumerate(ogrenci_numaralari): #Başlıkları ve içerikleri her öğrenci için ekler
    # Yeni tablonun başlangıç satırı (her öğrenci için 9 satır arayla olacak)
    base_row = 17 + (i * 8)  #8 satır atlar 
    sayfa.cell(row=base_row, column=10, value="Tablo 4")
    sayfa.cell(row=base_row, column=11, value=ogrenci_numarasi) #Öğrenci numarasını ekler
    başlık = ["Ders Çıktısı", "1. Başlık", "2. Başlık", "3. Başlık", "4. Başlık", "5. Başlık"]
    for idx, header in enumerate(başlık, start=base_row + 1): #Başlıkları hemen altına yazar
        sayfa.cell(row=idx, column=10, value=header) #10. sütun
    right_başlık = ["Ödev", "Quiz", "Vize", "Final", "Toplam", "Max", "% Başarı"]
    for idx, header in enumerate(right_başlık, start=11): 
        sayfa.cell(row=base_row + 1, column=idx, value=header) 
#------------------------------------------Tablo 3'teki verileri kullanarak tablo 4'teki değerleri oluşturma. Her bir sütun için ayrı ayrı.
row = 19  # Başlangıç satırı
row2 = 19
row3 = 20
row4 = 21
row5 = 22
row6 = 23
carpan_degeri1 = sayfa['R4'].value
carpan_degeri2 = sayfa['R5'].value
carpan_degeri3 = sayfa['R6'].value
carpan_degeri4 = sayfa['R7'].value
carpan_degeri5 = sayfa['R8'].value
while sayfa.cell(row=row, column=3).value is not None: #Öğrenci sütununda değer var mı kontrol eder
    not_degeri = sayfa.cell(row=row, column=3).value  #Öğrenci tablosundaki not değeri
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri1, (int, float)):
        sonuc = not_degeri * carpan_degeri1
        sayfa.cell(row=row2, column=11, value=sonuc)#K sütunu (11. sütun)
        row2 += 8  
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri2, (int, float)):
        sonuc = not_degeri * carpan_degeri2
        sayfa.cell(row=row3, column=11, value=sonuc)#K sütunu (11. sütun)
        row3 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri3, (int, float)):
        sonuc = not_degeri * carpan_degeri3
        sayfa.cell(row=row4, column=11, value=sonuc)#K sütunu (11. sütun)
        row4 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri4, (int, float)):
        sonuc = not_degeri * carpan_degeri4
        sayfa.cell(row=row5, column=11, value=sonuc)#K sütunu (11. sütun)
        row5 += 8
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri5, (int, float)):
        sonuc = not_degeri * carpan_degeri5
        sayfa.cell(row=row6, column=11, value=sonuc)#K sütunu (11. sütun)
        row6 += 8
    row += 1  #1 satır atla (notları okuma işlemi)
#------------------------------------------Tablo 3'teki verileri kullanarak tablo 4'teki değerleri oluşturma. Her bir sütun için ayrı ayrı.
row = 19  # Başlangıç satırı
row2 = 19
row3 = 20
row4 = 21
row5 = 22
row6 = 23
carpan_degeri1 = sayfa['S4'].value
carpan_degeri2 = sayfa['S5'].value
carpan_degeri3 = sayfa['S6'].value
carpan_degeri4 = sayfa['S7'].value
carpan_degeri5 = sayfa['S8'].value
while sayfa.cell(row=row, column=4).value is not None:  #Öğrenci sütununda değer var mı kontrol eder
    not_degeri = sayfa.cell(row=row, column=4).value  #Öğrenci tablosundaki not değeri
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri1, (int, float)):
        sonuc = not_degeri * carpan_degeri1
        sayfa.cell(row=row2, column=12, value=sonuc)#L sütunu (12. sütun)
        row2 += 8  
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri2, (int, float)):
        sonuc = not_degeri * carpan_degeri2
        sayfa.cell(row=row3, column=12, value=sonuc)#L sütunu (12. sütun)
        row3 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri3, (int, float)):
        sonuc = not_degeri * carpan_degeri3
        sayfa.cell(row=row4, column=12, value=sonuc)#L sütunu (12. sütun)
        row4 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri4, (int, float)):
        sonuc = not_degeri * carpan_degeri4
        sayfa.cell(row=row5, column=12, value=sonuc)#L sütunu (12. sütun)
        row5 += 8
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri5, (int, float)):
        sonuc = not_degeri * carpan_degeri5
        sayfa.cell(row=row6, column=12, value=sonuc)#L sütunu (12. sütun)
        row6 += 8
    row += 1  # 1 satır atla (notları okuma işlemi)
#------------------------------------------Tablo 3'teki verileri kullanarak tablo 4'teki değerleri oluşturma. Her bir sütun için ayrı ayrı.
row = 19  # Başlangıç satırı
row2 = 19
row3 = 20
row4 = 21
row5 = 22
row6 = 23
carpan_degeri1 = sayfa['T4'].value
carpan_degeri2 = sayfa['T5'].value
carpan_degeri3 = sayfa['T6'].value
carpan_degeri4 = sayfa['T7'].value
carpan_degeri5 = sayfa['T8'].value
while sayfa.cell(row=row, column=5).value is not None:  #Öğrenci sütununda değer var mı kontrol eder
    not_degeri = sayfa.cell(row=row, column=5).value  #Öğrenci tablosundaki not değeri
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri1, (int, float)):
        sonuc = not_degeri * carpan_degeri1
        sayfa.cell(row=row2, column=13, value=sonuc)#M sütunu (13. sütun)
        row2 += 8  
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri2, (int, float)):
        sonuc = not_degeri * carpan_degeri2
        sayfa.cell(row=row3, column=13, value=sonuc)#M sütunu (13. sütun)
        row3 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri3, (int, float)):
        sonuc = not_degeri * carpan_degeri3
        sayfa.cell(row=row4, column=13, value=sonuc)#M sütunu (13. sütun)
        row4 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri4, (int, float)):
        sonuc = not_degeri * carpan_degeri4
        sayfa.cell(row=row5, column=13, value=sonuc)#M sütunu (13. sütun)
        row5 += 8
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri5, (int, float)):
        sonuc = not_degeri * carpan_degeri5
        sayfa.cell(row=row6, column=13, value=sonuc)#M sütunu (13. sütun)
        row6 += 8
    row += 1  # 1 satır atla (notları okuma işlemi)
#------------------------------------------Tablo 3'teki verileri kullanarak tablo 4'teki değerleri oluşturma. Her bir sütun için ayrı ayrı.
row = 19  # Başlangıç satırı
row2 = 19
row3 = 20
row4 = 21
row5 = 22
row6 = 23
carpan_degeri1 = sayfa['U4'].value
carpan_degeri2 = sayfa['U5'].value
carpan_degeri3 = sayfa['U6'].value
carpan_degeri4 = sayfa['U7'].value
carpan_degeri5 = sayfa['U8'].value
while sayfa.cell(row=row, column=6).value is not None:  #Öğrenci sütununda değer var mı kontrol eder
    not_degeri = sayfa.cell(row=row, column=6).value  #Öğrenci tablosundaki not değeri
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri1, (int, float)):
        sonuc = not_degeri * carpan_degeri1
        sayfa.cell(row=row2, column=14, value=sonuc) #N sütunu (14. sütun)
        row2 += 8  
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri2, (int, float)):
        sonuc = not_degeri * carpan_degeri2
        sayfa.cell(row=row3, column=14, value=sonuc) #N sütunu (14. sütun)
        row3 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri3, (int, float)):
        sonuc = not_degeri * carpan_degeri3
        sayfa.cell(row=row4, column=14, value=sonuc) #N sütunu (14. sütun)
        row4 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri4, (int, float)):
        sonuc = not_degeri * carpan_degeri4
        sayfa.cell(row=row5, column=14, value=sonuc) #N sütunu (14. sütun)
        row5 += 8
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri5, (int, float)):
        sonuc = not_degeri * carpan_degeri5
        sayfa.cell(row=row6, column=14, value=sonuc) #N sütunu (14. sütun)
        row6 += 8
    row += 1  # 1 satır atla (notları okuma işlemi)
#------------------------------------------Tablo 4'teki verileri kullanarak tablo 4'teki TOPLAM değerini oluşturma.
start_row = 19  #İşlem yapılacak ilk satır
row_offset = 0  #Satır kaydırması için başlangıç değeri
value_count = 0
while sayfa.cell(row=start_row + value_count, column=3).value is not None: #Öğrenci sütununda kaç değer olduğunu kontrol eder
    value_count += 1  #Öğrenci sütununda değer varsa 1 artar
for i in range(value_count): #Her öğrenci için
    for row in range(start_row + row_offset, start_row + row_offset + 5):
        toplam = 0
        for col in range(11, 15):
            value = sayfa.cell(row=row, column=col).value
            if isinstance(value, (int, float)): 
                toplam += value
        sayfa.cell(row=row, column=15, value=toplam) #Toplam sütununa değeri yazar
    row_offset += 8  
#------------------------------------------Tablo 3'teki verileri kullanarak tablo 4'teki değerleri oluşturma. Her bir sütun için ayrı ayrı.
row = 19  # Başlangıç satırı
row2 = 19
row3 = 20
row4 = 21
row5 = 22
row6 = 23
carpan_degeri1 = sayfa['V4'].value
carpan_degeri2 = sayfa['V5'].value
carpan_degeri3 = sayfa['V6'].value
carpan_degeri4 = sayfa['V7'].value
carpan_degeri5 = sayfa['V8'].value
while sayfa.cell(row=row, column=7).value is not None: #Öğrenci sütununda değer var mı kontrol eder
    not_degeri = sayfa.cell(row=row, column=7).value  #Öğrenci tablosundaki not değeri
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri1, (int, float)):
        sonuc = 100 * carpan_degeri1
        sayfa.cell(row=row2, column=16, value=sonuc) #O sütunu (16. sütun)
        row2 += 8  
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri2, (int, float)):
        sonuc = 100 * carpan_degeri2
        sayfa.cell(row=row3, column=16, value=sonuc) #O sütunu (16. sütun)
        row3 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri3, (int, float)):
        sonuc = 100 * carpan_degeri3
        sayfa.cell(row=row4, column=16, value=sonuc) #O sütunu (16. sütun)
        row4 += 8 
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri4, (int, float)):
        sonuc = 100 * carpan_degeri4
        sayfa.cell(row=row5, column=16, value=sonuc) #O sütunu (16. sütun)
        row5 += 8
    if isinstance(not_degeri, (int, float)) and isinstance(carpan_degeri5, (int, float)):
        sonuc = 100 * carpan_degeri5
        sayfa.cell(row=row6, column=16, value=sonuc) #O sütunu (16. sütun)
        row6 += 8
    row += 1  # 1 satır atla (notları okuma işlemi)

start_row = 19  # İşlem yapılacak ilk satır
row_offset = 0  # Satır kaydırması için başlangıç değeri
value_count = 0
while sayfa.cell(row=start_row + value_count, column=3).value is not None:
    value_count += 1 
for i in range(value_count):  
    for row in range(start_row + row_offset, start_row + row_offset + 5):
        #O sütunundaki değeri 100 ile çarpıp P sütunundaki değere böler
        o_value = sayfa.cell(row=row, column=15).value #O sütunundaki değer
        p_value = sayfa.cell(row=row, column=16).value #P sütunundaki değer
        if isinstance(o_value, (int, float)) and isinstance(p_value, (int, float)) and p_value != 0:
            sonuc = (o_value * 100) / p_value  #O'yu 100 ile çarp, P'ye böl
            sonuc = round(sonuc, 1)  #Virgülden sonra 1 basamağa yuvarla
            sayfa.cell(row=row, column=17, value=sonuc).font = Font(color="FF0000") #17. sütununa sonucu yaz(kırmızı renk)
        else:
            sayfa.cell(row=row, column=17, value=None)
    row_offset += 8  
#-------------------------------------------------------Tablo 5'i oluşturma               
ogrenci_numaralari = []
row = 19 #B19 hücresinden başlayacağız
#Öğrenci numaralarının hepsini alalım
while sayfa.cell(row=row, column=2).value is not None: #Sütunda öğrenci numarası varsa
    ogrenci_numaralari.append(sayfa.cell(row=row, column=2).value)
    row += 1  #Sonraki satıra geç
#Başlıkları ve içerikleri her öğrenci için ekler
for i, ogrenci_numarasi in enumerate(ogrenci_numaralari):
    base_row = 17 + (i * 13) #13 aralık bırakır
    sayfa.cell(row=base_row, column=19, value="Tablo 5")
    sayfa.cell(row=base_row, column=20, value=ogrenci_numarasi) #Öğrenci numarasını ekler
    sayfa.merge_cells(start_row=base_row, start_column=21, end_row=base_row, end_column=24)
    sayfa.cell(row=base_row, column=21, value="Ders Çıktısı")
    başlık = ["P Çıktısı","Prgm 1", "Prgm 2", "Prgm 3", "Prgm 4", "Prgm 5", "Prgm 6", "Prgm 7", "Prgm 8", "Prgm 9", "Prgm 10"]
    for idx, header in enumerate(başlık, start=base_row + 1):  #Başlıkları hemen altına değerleri yazalım
        sayfa.cell(row=idx, column=19, value=header)#19. sütun
    right_başlık = ["Başarı Oranı"]
    for idx, header in enumerate(right_başlık, start=25):  
        sayfa.cell(row=base_row + 1, column=idx, value=header) 
#------------------------------Tablo 4'ten alınan değerlerle tablo 5'te işlem(her sütun için)
start_row = 18  #18. satır
offset = 13  #13 satır aralıklarla yazacağız
row_offset = 19  #Başlangıç 19. satırdan
while True:
    sayisi = sayfa[f'Q{row_offset}'].value #Q hücresinden alınan sayıyı alalım
    if isinstance(sayisi, (int, float)):
        sayfa.cell(row=start_row, column=20, value=sayisi) #T sütununa yazma
        start_row += offset  # 13 satır kaydırıyoruz
    else: 
        break
    row_offset += 8 #8 satır aşağıya geçiyoruz
#------------------------------Tablo 4'ten alınan değerlerle tablo 5'te işlem(her sütun için)
start_row = 18  
offset = 13  
row_offset = 20  
while True:
    sayisi = sayfa[f'Q{row_offset}'].value #Q hücresinden alınan sayıyı alalım
    if isinstance(sayisi, (int, float)):
        sayfa.cell(row=start_row, column=21, value=sayisi)  #U sütununa yazma
        start_row += offset  # 13 satır kaydırıyoruz
    else:
        break
    row_offset += 8
#------------------------------Tablo 4'ten alınan değerlerle tablo 5'te işlem(her sütun için)
start_row = 18  # V18 hücresine yazacağız
offset = 13  # 13 satır aralıklarla yazacağız
row_offset = 21  # Başlangıç Q21'den alıyoruz
while True:
    sayisi = sayfa[f'Q{row_offset}'].value # Qhücresinden alınan sayıyı alalım
    if isinstance(sayisi, (int, float)):
        sayfa.cell(row=start_row, column=22, value=sayisi) #V sütununa yazma
        start_row += offset  # 13 satır kaydırıyoruz
    else:
        break
    row_offset += 8
#------------------------------Tablo 4'ten alınan değerlerle tablo 5'te işlem(her sütun için)
start_row = 18  # W18 hücresine yazacağız
offset = 13  
row_offset = 22 
while True:
    sayisi = sayfa[f'Q{row_offset}'].value 
    if isinstance(sayisi, (int, float)):
        sayfa.cell(row=start_row, column=23, value=sayisi)  
        start_row += offset  
    else:
        break
    row_offset += 8
#------------------------------Tablo 4'ten alınan değerlerle tablo 5'te işlem(her sütun için)
start_row = 18  # X18 hücresine yazacağız
offset = 13  
row_offset = 23  
while True:
    sayisi = sayfa[f'Q{row_offset}'].value 
    if isinstance(sayisi, (int, float)):
        sayfa.cell(row=start_row, column=24, value=sayisi)  
        start_row += offset  
    else:
        break
    row_offset += 8
#------------------------------Tablo 5 için çarpım işlmei(her satır için ayrı)
current_row = 18  # İlk işlem T18'den başlayacak
carpanlar = [sayfa[f'C{i}'].value for i in range(4, 14)] #C4:C13 arası değerler
while True:
    deger = sayfa[f'T{current_row}'].value #T hücresindeki değeri al
    if deger is None: #Değerler bitince dur
        break
    for i, carpan in enumerate(carpanlar):
        if isinstance(carpan, (int, float)) and isinstance(deger, (int, float)):
            sonuc = carpan * deger
            sayfa[f'T{current_row + i + 1}'] = sonuc  #Sonuçları bir satır aşağıya yaz
    current_row += 13
#------------------------------Tablo 5 için çarpım işlmei(her satır için ayrı)
current_row = 18  #İlk işlem U18'den başlayacak
carpanlar = [sayfa[f'D{i}'].value for i in range(4, 14)]  #D4:D13 arası 
while True:
    deger = sayfa[f'U{current_row}'].value #U hücresindeki değeri al
    if deger is None:
        break
    for i, carpan in enumerate(carpanlar): #Çarpma işlemi yap ve sonucu yaz
        if isinstance(carpan, (int, float)) and isinstance(deger, (int, float)):
            sonuc = carpan * deger
            sayfa[f'U{current_row + i + 1}'] = sonuc  #Sonuçları bir satır aşağıya yaz(U'ya)
    current_row += 13
    #------------------------------Tablo 5 için çarpım işlmei(her satır için ayrı)
current_row = 18 
carpanlar = [sayfa[f'E{i}'].value for i in range(4, 14)] # D4:D13 arası çarpanlar
while True:
    deger = sayfa[f'V{current_row}'].value
    if deger is None:
        break
    for i, carpan in enumerate(carpanlar): #Çarpma işlemi yap ve sonucu yaz
        if isinstance(carpan, (int, float)) and isinstance(deger, (int, float)):
            sonuc = carpan * deger
            sayfa[f'V{current_row + i + 1}'] = sonuc  #Sonuçları bir satır aşağıya yaz
    current_row += 13
#------------------------------Tablo 5 için çarpım işlmei(her satır için ayrı)
current_row = 18  # İlk işlem U18'den başlayacak
carpanlar = [sayfa[f'F{i}'].value for i in range(4, 14)]  #D4:D13 arası çarpanlar
while True:
    deger = sayfa[f'W{current_row}'].value
    if deger is None:
        break
    for i, carpan in enumerate(carpanlar):
        if isinstance(carpan, (int, float)) and isinstance(deger, (int, float)):
            sonuc = carpan * deger
            sayfa[f'W{current_row + i + 1}'] = sonuc  #Sonuçları bir satır aşağıya yaz
    current_row += 13
#------------------------------Tablo 5 için çarpım işlmei(her satır için ayrı)
current_row = 18  
carpanlar = [sayfa[f'G{i}'].value for i in range(4, 14)]  #D4:D13 arası çarpanlar
while True:
    deger = sayfa[f'X{current_row}'].value
    if deger is None:
        break
    for i, carpan in enumerate(carpanlar):
        if isinstance(carpan, (int, float)) and isinstance(deger, (int, float)):
            sonuc = carpan * deger
            sayfa[f'X{current_row + i + 1}'] = sonuc 
    current_row += 13

# Başlangıç değerleri
initial_row = 19  # Başlangıç satırı
h_start_row = 4  # H sütunundaki başlangıç hücresi
step = 13  # 13 satır aşağı inme
while True:
    current_row = initial_row
    h_row = h_start_row
    while True:  #H sütunu bitene kadar işle
        h_degeri = sayfa[f'H{h_row}'].value #H sütunundaki değeri al
        if h_degeri is None: #H sütunu bitince durdurur
            break
        t_value = sayfa[f'T{current_row}'].value # Verileri T, U, V, W, X sütunlarından al
        u_value = sayfa[f'U{current_row}'].value
        v_value = sayfa[f'V{current_row}'].value
        w_value = sayfa[f'W{current_row}'].value
        x_value = sayfa[f'X{current_row}'].value
        if any(val is None for val in [t_value, u_value, v_value, w_value, x_value]):
            break
        toplam = t_value + u_value + v_value + w_value + x_value # T, U, V, W, X sütunlarındaki değerleri topla ve ortalama al
        ortalama = toplam / 5
        if isinstance(h_degeri, (int, float)) and h_degeri != 0: #Ortalamayı H sütunundaki değere böl
            sonuc_y = ortalama / h_degeri
            y_cell = sayfa[f'Y{current_row}']
            y_cell.value = round(sonuc_y, 1)  #Yuvarlanmış sonuç
            y_cell.font = Font(color="FF0000") #Kırmızı font
        h_row += 1
        current_row += 1
    next_row = initial_row + step
    if sayfa[f'T{next_row}'].value is None:
        break
    initial_row = next_row #Aşağı in ve işlemi tekrarla

################################################ Hücreler için otomatik boyutlandırma
def auto_resize_cells(sayfa):
    for col in range(1, sayfa.max_column + 1): 
        max_length = 0
        column = get_column_letter(col)
        # Her sütun için en uzun metni bul
        for row in range(1, sayfa.max_row + 1):
            cell_value = str(sayfa.cell(row=row, column=col).value) if sayfa.cell(row=row, column=col).value else ""     # Hücreler için otomatik boyutlandırma
            max_length = max(max_length, len(cell_value))    
        sayfa.column_dimensions[column].width = max_length
    for row in range(1, sayfa.max_row + 1):
        max_row_height = 0
        for col in range(1, sayfa.max_column + 1):
            cell_value = str(sayfa.cell(row=row, column=col).value) if sayfa.cell(row=row, column=col).value else ""     # Hücreler için otomatik boyutlandırma
            # Satırın yüksekliğini, metnin uzunluğuna göre ayarlıyoruz
            max_row_height = max(max_row_height, len(cell_value) // 50 + 1)  # 50 karakteri bir satır olarak kabul et    # Hücreler için otomatik boyutlandırma
        # Satır yüksekliğini ayarla
        sayfa.row_dimensions[row].height = max_row_height * 15  # Her satırda 15px yükseklik artırıyoruz
    # Hücrelerde metni ortalamak ve sarmalamak
    for row in range(1, sayfa.max_row + 1):
        for col in range(1, sayfa.max_column + 1):
            sayfa.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) # Hücreler için otomatik boyutlandırma
# Otomatik ayarlamaları yapalım
auto_resize_cells(sayfa)
################################################ Hücreler için otomatik boyutlandırma

def adjust_column_width(sayfa): #Sütun genişliklerini ayarlamak için
    columns_to_adjust = ["G","P","K","L","O","M","N",'R','S','T','U',"V","W","X"] #"Ödev", "Quiz", "Vize", "Final" gibi sütunlarının genişliğini arttıralım
    new_width = 10  #Yeni sütun genişliği, ihtiyaca göre ayarlanabilir
    for column in columns_to_adjust:
        sayfa.column_dimensions[column].width = new_width  #Yeni genişlik ayarı
adjust_column_width(sayfa) #Sütunları için otomatik boyutlandırma



ç_sayfası.save(file_path)  #Değişiklikleri kaydet